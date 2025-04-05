from django.shortcuts import render
from django.utils import timezone
from django.views.decorators.csrf import csrf_exempt
from django.http import JsonResponse
from rest_framework.decorators import api_view
from rest_framework.response import Response
from rest_framework import status

import qrcode
import io
import base64
import json
import os
from openpyxl import Workbook, load_workbook

from .models import Faculty, AttendanceSession, AttendanceRecord
from .decorators import login_required_or_pin_check

# ---------- QR Code API ----------
@api_view(['POST'])
def create_attendance_qr(request):
    session = AttendanceSession.objects.create()
    session_url = f"https://qr-attendance-hxvv.onrender.com/faculty-login/{session.session_id}/"

    qr = qrcode.make(session_url)
    buffer = io.BytesIO()
    qr.save(buffer, format='PNG')
    qr_base64 = base64.b64encode(buffer.getvalue()).decode()

    return Response({
        "session_id": str(session.session_id),
        "qr_code": f"data:image/png;base64,{qr_base64}",
        "url": session_url
    })

# ---------- For Manual API testing ----------
@csrf_exempt
def mark_attendance(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            pin = data.get('pin')

            if pin == "1234":
                return JsonResponse({'status': 'success', 'message': 'Attendance marked successfully!'})
            else:
                return JsonResponse({'status': 'fail', 'message': 'Invalid PIN'}, status=400)
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=500)

    return JsonResponse({'error': 'Invalid request method'}, status=405)

# ---------- API for submission ----------
@login_required_or_pin_check
@api_view(['POST'])
def submit_attendance(request):
    name = request.data.get('name')
    email = request.data.get('email')
    pin = request.data.get('pin')
    session_id = request.data.get('session_id')

    try:
        faculty = Faculty.objects.get(name=name, email=email, pin=pin)
    except Faculty.DoesNotExist:
        return Response({"error": "Invalid name, email or PIN"}, status=401)

    try:
        session = AttendanceSession.objects.get(session_id=session_id)
    except AttendanceSession.DoesNotExist:
        return Response({"error": "Invalid session ID"}, status=404)

    now = timezone.now()
    AttendanceRecord.objects.create(faculty=faculty, session=session, timestamp=now, action="login")

    return Response({
        "message": "Attendance marked successfully!",
        "login_time": now.strftime("%Y-%m-%d %H:%M:%S")
    })

# ---------- Login/Logout Web Form ----------
def attendance_form(request, session_id):
    if request.method == 'POST':
        name = request.POST.get('name')
        pin = request.POST.get('pin')
        action = request.POST.get('action')

        try:
            faculty = Faculty.objects.get(name=name, pin=pin)
            session = AttendanceSession.objects.get(session_id=session_id, active=True)

            AttendanceRecord.objects.create(
                session=session,
                faculty=faculty,
                action=action,
                timestamp=timezone.now()
            )

            # Save to Excel
            file_path = "attendance_records.xlsx"
            if os.path.exists(file_path):
                wb = load_workbook(file_path)
                ws = wb.active
            else:
                wb = Workbook()
                ws = wb.active
                ws.append(['Name', 'PIN', 'Action', 'Timestamp', 'Session ID'])

            now_str = timezone.now().strftime("%Y-%m-%d %H:%M:%S")
            ws.append([faculty.name, faculty.pin, action, now_str, str(session.session_id)])
            wb.save(file_path)

            return render(request, 'attendance/success.html', {
                'name': faculty.name,
                'action': action,
                'time': now_str
            })
        except Faculty.DoesNotExist:
            return render(request, 'attendance/form.html', {'error': 'Invalid Name or PIN'})
        except AttendanceSession.DoesNotExist:
            return render(request, 'attendance/form.html', {'error': 'Invalid or inactive session ID'})

    return render(request, 'attendance/form.html')
